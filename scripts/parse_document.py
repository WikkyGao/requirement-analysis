#!/usr/bin/env python3
"""
parse_document.py - 多格式文档解析为统一中间 JSON
接收 stdin JSON: {"path": "...", "format": "markdown|docx|xlsx|xmind|mm"}
输出 stdout JSON: {"status": "ok", "data": {...}} 或 {"status": "error", "message": "..."}
"""
import json
import os
import re
import sys
import xml.etree.ElementTree as ET
from zipfile import ZipFile
from io import StringIO
from pathlib import Path


ALLOWED_FORMATS = frozenset({'markdown', 'docx', 'xlsx', 'xmind', 'mm'})


def validate_input(data: dict) -> None:
    if 'path' not in data or not isinstance(data['path'], str):
        raise ValueError('"path" is required and must be a string')
    if 'format' not in data or data['format'] not in ALLOWED_FORMATS:
        raise ValueError(f'"format" must be one of: {", ".join(sorted(ALLOWED_FORMATS))}')
    if not os.path.isfile(data['path']):
        raise ValueError(f'File not found: {data["path"]}')


def parse_markdown(path: str) -> dict:
    with open(path, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.split('\n')
    result = {
        'title': '',
        'modules': [],
        'raw_content': content
    }

    current_module = None
    current_feature = None
    in_table = False
    table_headers = []
    table_rows = []

    for line in lines:
        # Title: first # heading
        if line.startswith('# ') and not result['title']:
            result['title'] = line[2:].strip()
            continue

        # Module: ## heading
        if line.startswith('## '):
            if current_feature and current_module:
                current_module['features'].append(current_feature)
                current_feature = None
            if current_module:
                result['modules'].append(current_module)
            current_module = {
                'name': line[3:].strip(),
                'features': []
            }
            continue

        # Feature: ### heading
        if line.startswith('### '):
            if current_feature and current_module:
                current_module['features'].append(current_feature)
            current_feature = {
                'name': line[4:].strip(),
                'description': '',
                'fields': [],
                'details': []
            }
            continue

        # Table detection
        if line.startswith('|') and line.endswith('|'):
            if not in_table:
                in_table = True
                parts = [p.strip() for p in line.split('|')[1:-1]]
                table_headers = parts
                table_rows = []
            else:
                # Check if it's a separator row (|---|)
                if re.match(r'^\|[\s\-:]+\|$', line):
                    continue
                parts = [p.strip() for p in line.split('|')[1:-1]]
                if parts and len(parts) == len(table_headers):
                    table_rows.append(dict(zip(table_headers, parts)))
            continue
        else:
            if in_table:
                # Store collected table data
                in_table = False
                if current_feature:
                    current_feature['fields'] = table_rows
                elif current_module:
                    pass  # module-level tables
                table_headers = []
                table_rows = []

        # List items as details
        if line.strip().startswith('- ') and current_feature:
            current_feature['details'].append(line.strip()[2:])
            continue

        # Regular text as description
        if line.strip() and current_feature:
            if current_feature['description']:
                current_feature['description'] += ' ' + line.strip()
            else:
                current_feature['description'] = line.strip()

    # Flush last feature and module
    if current_feature and current_module:
        current_module['features'].append(current_feature)
    if current_module:
        result['modules'].append(current_module)

    return result


def parse_docx(path: str) -> dict:
    """Parse .docx file - requires python-docx"""
    try:
        import docx
    except ImportError:
        return {
            'status': 'error',
            'message': 'python-docx not installed. Run: pip install python-docx'
        }

    doc = docx.Document(path)
    text_lines = [p.text for p in doc.paragraphs]

    # Build the same structure as markdown parser using heading levels
    content = '\n'.join(text_lines)
    # Write to temp and use markdown parser on the text
    # (docx doesn't have markdown tables, just text)

    return parse_markdown_text(content)


def parse_markdown_text(content: str) -> dict:
    """Parse text content as if it were markdown (for docx fallback)."""
    lines = content.split('\n')
    result = {'title': '', 'modules': [], 'raw_content': content}
    current_module = None
    current_feature = None

    for line in lines:
        stripped = line.strip()
        if stripped.startswith('# ') and not result['title']:
            result['title'] = stripped[2:].strip()
            continue
        if stripped.startswith('## '):
            if current_feature and current_module:
                current_module['features'].append(current_feature)
                current_feature = None
            if current_module:
                result['modules'].append(current_module)
            current_module = {'name': stripped[3:].strip(), 'features': []}
            continue
        if stripped.startswith('### '):
            if current_feature and current_module:
                current_module['features'].append(current_feature)
            current_feature = {'name': stripped[4:].strip(), 'description': '', 'fields': [], 'details': []}
            continue
        if stripped and current_feature:
            if current_feature['description']:
                current_feature['description'] += ' ' + stripped
            else:
                current_feature['description'] = stripped

    if current_feature and current_module:
        current_module['features'].append(current_feature)
    if current_module:
        result['modules'].append(current_module)
    return result


def parse_excel(path: str) -> dict:
    """Parse .xlsx file - requires openpyxl"""
    try:
        import openpyxl
    except ImportError:
        return {'status': 'error', 'message': 'openpyxl not installed. Run: pip install openpyxl'}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {'sheets': [], 'title': os.path.basename(path)}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h is not None else '' for h in rows[0]]
        data_rows = []
        for row in rows[1:]:
            row_data = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = str(val) if val is not None else ''
            if any(v for v in row_data.values()):
                data_rows.append(row_data)

        result['sheets'].append({
            'name': sheet_name,
            'headers': headers,
            'rows': data_rows
        })

    wb.close()
    return result


def parse_xmind(path: str) -> dict:
    """Parse .xmind file - extract content.xml from zip, parse XML tree."""
    try:
        with ZipFile(path, 'r') as z:
            if 'content.xml' in z.namelist():
                xml_content = z.read('content.xml')
            else:
                # Try to find any xml file
                xml_files = [n for n in z.namelist() if n.endswith('.xml')]
                if not xml_files:
                    return {'status': 'error', 'message': 'No content.xml found in .xmind file'}
                xml_content = z.read(xml_files[0])
    except Exception as e:
        return {'status': 'error', 'message': f'Failed to read .xmind file: {str(e)}'}

    return parse_xml_mindmap(xml_content)


def parse_freemind(path: str) -> dict:
    """Parse .mm FreeMind file."""
    try:
        with open(path, 'r', encoding='utf-8') as f:
            xml_content = f.read()
    except Exception as e:
        return {'status': 'error', 'message': f'Failed to read .mm file: {str(e)}'}

    return parse_xml_mindmap(xml_content.encode('utf-8'))


def parse_xml_mindmap(xml_data: bytes) -> dict:
    """Parse XML mind map data into structured modules/features."""
    try:
        root = ET.fromstring(xml_data)
    except ET.ParseError as e:
        return {'status': 'error', 'message': f'XML parse error: {str(e)}'}

    # Namespace handling
    ns = {}
    # Try to find namespaces
    for m in re.finditer(r'xmlns:?(\w*)=["\']([^"\']+)["\']', xml_data.decode('utf-8', errors='ignore')):
        prefix, uri = m.group(1), m.group(2)
        ns[prefix or 'default'] = uri

    def get_text(element):
        """Extract text from an XML element."""
        texts = [element.text or '']
        for child in element:
            if child.tag.endswith('}text') or (child.tag.endswith('}html') and child.text):
                texts.append(child.text or '')
        return ''.join(texts).strip()

    def extract_topics(element, depth=0):
        """Recursively extract topics from mind map."""
        items = []
        for child in element:
            tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
            if tag in ('topic', 'node'):
                text = get_text(child)
                children = extract_topics(child, depth + 1)
                items.append({'name': text, 'children': children})
        return items

    # Find the root topic
    topics = []
    for child in root.iter():
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if tag in ('topic', 'node') and child.text and child.text.strip():
            text = get_text(child)
            children = extract_topics(child, 1)
            if children or text:
                topics.append({'name': text, 'children': children})
            break

    # Convert to module/feature structure
    result = {'title': '', 'modules': []}
    if topics:
        root_topic = topics[0]
        result['title'] = root_topic['name']

        for module_node in root_topic['children']:
            module = {'name': module_node['name'], 'features': []}
            for feature_node in module_node['children']:
                feature = {
                    'name': feature_node['name'],
                    'description': '',
                    'fields': [],
                    'details': [c['name'] for c in feature_node['children']]
                }
                module['features'].append(feature)
            result['modules'].append(module)

    return result


def main():
    try:
        raw = sys.stdin.read()
        if not raw.strip():
            print(json.dumps({'status': 'error', 'message': 'Empty stdin input'}))
            sys.exit(1)

        data = json.loads(raw)
        validate_input(data)

        fmt = data['format']
        path = data['path']

        parsers = {
            'markdown': parse_markdown,
            'docx': parse_docx,
            'xlsx': parse_excel,
            'xmind': parse_xmind,
            'mm': parse_freemind,
        }

        parser = parsers[fmt]
        result = parser(path)

        if isinstance(result, dict) and result.get('status') == 'error':
            print(json.dumps(result))
            sys.exit(1)

        output = {
            'status': 'ok',
            'format': fmt,
            'path': path,
            'data': result
        }
        print(json.dumps(output, ensure_ascii=False))

    except json.JSONDecodeError as e:
        print(json.dumps({'status': 'error', 'message': f'Invalid JSON input: {str(e)}'}))
        sys.exit(1)
    except ValueError as e:
        print(json.dumps({'status': 'error', 'message': str(e)}))
        sys.exit(1)
    except Exception as e:
        print(json.dumps({'status': 'error', 'message': f'Unexpected error: {str(e)}'}))
        sys.exit(1)


if __name__ == '__main__':
    main()
